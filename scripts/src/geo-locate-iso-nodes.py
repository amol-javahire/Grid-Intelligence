#!/usr/bin/env python3
"""
geo-locate-iso-nodes.py — locate ISO settlement points, with honest provenance.

Market-parameterised so the same code serves CAISO now, re-runs ERCOT with the
precision columns added, and covers PJM when pjm_nodal_da_rt_hourly is built.

    --market CAISO          (required)
    --apply                 write to DB; without it this is a DRY RUN
    --limit N               process only N nodes, for a quick look
    --skip-usgs             skip USPVDB/USWTDB (offline or endpoint changed)

PHASES, best evidence first. A node is only ever UPGRADED — a later phase
never overwrites a more precise earlier result.

  1. EIA-860 exact LMP node match      precision=facility  conf 0.90
     3_1_Generator "RTO/ISO LMP Node Designation" IS the settlement point name.
     Coordinates come from 2___Plant via Plant Code.

  2. USPVDB — solar                     precision=exact     conf 0.98
  3. USWTDB — wind                      precision=exact     conf 0.98
     USGS databases with imagery-verified array boundaries and individual
     turbine positions (~10m). Joined on EIA plant code, so they UPGRADE the
     coordinate of a node already identified in phase 1 rather than matching
     names independently. Strictly better than EIA-860's reported point.

  4. EIA-860 fuzzy plant-name match     precision=facility  conf 0.50-0.75
     Only above a similarity floor, and the score is stored.

  NO COUNTY-CENTROID PHASE. ercot_node_locations has 134 of 439 rows as county
  centroids sitting in the same latitude column as surveyed coordinates. Texas
  counties are 50-100km across. An unmatched node is left NULL here: absent is
  honest, a 50km guess presented as a location is not.

WHY PROVENANCE COLUMNS AND NOT JUST A COORDINATE
  Basis risk keys off the POINT OF INTERCONNECTION; resource quality keys off
  the SITE. Public ISO queues never publish parcel coordinates, so any queue-
  derived point is a POI or a county centroid. Storing both in one unlabelled
  lat/lon column makes those indistinguishable, which is how a +/-50km guess
  ends up driving a "nearest node" calculation.

Run:
  cd ~/grid-intelligence && set -a; source .env; set +a
  artifacts/pypsa-engine/.venv/bin/python3 scripts/src/geo-locate-iso-nodes.py --market CAISO
  artifacts/pypsa-engine/.venv/bin/python3 scripts/src/geo-locate-iso-nodes.py --market CAISO --apply
"""

import argparse
import io
import math
import os
import re
import sys
import zipfile
from collections import Counter
from dataclasses import dataclass, field

import psycopg2
import psycopg2.extras
import requests

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required:  pip install openpyxl")

try:
    from rapidfuzz import fuzz
except ImportError:
    fuzz = None   # phase 4 degrades to exact-normalised matching only

DB_URL = os.environ.get("DATABASE_URL")
if not DB_URL:
    sys.exit("DATABASE_URL not set")

HERE = os.path.dirname(os.path.abspath(__file__))
EIA860_ZIP = os.path.join(HERE, "../../attached_assets/eia8602024_1777780153233.zip")

# USGS Energy and Environmental Resources Science Center APIs.
# VERIFY THESE if a phase returns zero rows — USGS has changed paths before.
# Both phases fail soft: a bad endpoint downgrades precision from 'exact' to
# 'facility', it does not abort the run or fabricate anything.
USPVDB_URL = "https://eersc.usgs.gov/api/uspvdb/v1/graphql"
USWTDB_URL = "https://eersc.usgs.gov/api/uswtdb/v1/turbines"

MARKETS = {
    "CAISO": {
        "state": "CA",
        "stats_table": "caiso_node_stats",
        "loc_table": "caiso_node_locations",
        "node_col": "node",
    },
    "ERCOT": {
        "state": "TX",
        "stats_table": "ercot_node_stats",
        "loc_table": "ercot_node_locations",
        "node_col": "node",
    },
}

# DLAP load centres — the same four points used for iso_hourly_temps, so
# temperature and node zone assignment cannot disagree with each other.
# Voronoi assignment: CAISO does not publish DLAP boundary geometry.
CAISO_DLAP_CENTRES = {
    "PGAE": (37.96, -121.29),   # Stockton
    "SCE":  (33.95, -117.40),   # Inland Empire
    "SDGE": (32.72, -117.16),   # San Diego
    "VEA":  (36.21, -115.98),   # Pahrump NV
}


@dataclass
class NodeLoc:
    node_name: str
    node_type: str | None = None
    latitude: float | None = None
    longitude: float | None = None
    precision: str = "unknown"
    method: str | None = None
    source: str | None = None
    confidence: float | None = None
    match_score: float | None = None
    eia_plant_code: int | None = None
    eia_plant_name: str | None = None
    technology: str | None = None
    avg_da_price: float | None = None
    months_available: int | None = None

    # Ordering used to decide whether a later phase may overwrite an earlier one.
    RANK = {"unknown": 0, "zone": 1, "county": 2, "city": 3,
            "poi": 4, "facility": 5, "exact": 6}

    def would_upgrade_to(self, precision: str) -> bool:
        """True if `precision` is strictly better than what this node already has."""
        return self.RANK[precision] > self.RANK[self.precision]


def norm(s: str) -> str:
    """Normalise a name for comparison: upper, alphanumeric only."""
    return re.sub(r"[^A-Z0-9]", "", (s or "").upper())


def node_prefix(node: str) -> str:
    """
    CAISO resource nodes look like VOLTA2_7_N002 or BALCH1_7_B2 — the plant
    identity is the token before the voltage segment. ERCOT settlement points
    use a similar leading-token convention.
    """
    return norm(node.split("_")[0])


def haversine_km(a: tuple[float, float], b: tuple[float, float]) -> float:
    R = 6371.0
    dlat, dlon = math.radians(b[0] - a[0]), math.radians(b[1] - a[1])
    h = (math.sin(dlat / 2) ** 2
         + math.cos(math.radians(a[0])) * math.cos(math.radians(b[0]))
         * math.sin(dlon / 2) ** 2)
    return 2 * R * math.asin(math.sqrt(h))


def caiso_zone_from_latlon(lat: float, lon: float) -> str:
    """
    APPROXIMATE. Path 15 sits near 37.0N (Los Banos) and Path 26 near 35.0N
    (Midway-Vincent). The real boundaries are electrical, not latitudinal, so
    nodes near a band edge may be assigned to the wrong zone. Recorded as
    approximate in the column comment; do not present as authoritative.
    """
    if lat >= 37.0:
        return "NP15"
    if lat >= 35.0:
        return "ZP26"
    return "SP15"


def nearest_dlap(lat: float, lon: float) -> str:
    return min(CAISO_DLAP_CENTRES,
               key=lambda d: haversine_km((lat, lon), CAISO_DLAP_CENTRES[d]))


# ── EIA-860 ─────────────────────────────────────────────────────────────────

def load_eia860(state: str):
    """Return (plants_by_code, generators) for one state."""
    if not os.path.exists(EIA860_ZIP):
        sys.exit(f"EIA-860 archive not found at {EIA860_ZIP}")
    z = zipfile.ZipFile(EIA860_ZIP)

    def sheet(fname: str, sheetname: str | None = None):
        wb = openpyxl.load_workbook(io.BytesIO(z.read(fname)), read_only=True, data_only=True)
        ws = wb[sheetname] if sheetname else wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        next(it)                                     # title row
        hdr = [str(h).strip() if h else "" for h in next(it)]
        return {h: i for i, h in enumerate(hdr)}, it

    H, rows = sheet("2___Plant_Y2024.xlsx")
    plants: dict[int, dict] = {}
    for r in rows:
        if not r or r[H["Plant Code"]] in (None, ""):
            continue
        if str(r[H["State"]]).strip() != state:
            continue
        try:
            lat, lon = float(r[H["Latitude"]]), float(r[H["Longitude"]])
        except (TypeError, ValueError):
            continue
        plants[int(r[H["Plant Code"]])] = {
            "name": str(r[H["Plant Name"]]).strip(),
            "lat": lat, "lon": lon,
            "county": str(r[H["County"]] or "").strip(),
        }

    G, grows = sheet("3_1_Generator_Y2024.xlsx", "Operable")
    gens = []
    for r in grows:
        if not r or r[G["Plant Code"]] in (None, ""):
            continue
        if str(r[G["State"]]).strip() != state:
            continue
        gens.append({
            "plant_code": int(r[G["Plant Code"]]),
            "plant_name": str(r[G["Plant Name"]]).strip(),
            "lmp_node": str(r[G["RTO/ISO LMP Node Designation"]] or "").strip(),
            "technology": str(r[G["Technology"]] or "").strip(),
        })

    print(f"  EIA-860 {state}: {len(plants):,} plants with coordinates, "
          f"{len(gens):,} operable generators, "
          f"{sum(1 for g in gens if g['lmp_node']):,} carrying an LMP node")
    return plants, gens


# ── USGS imagery-verified upgrades ──────────────────────────────────────────

def fetch_uspvdb() -> dict[int, tuple[float, float]]:
    """EIA plant code → imagery-verified solar array centroid."""
    q = {"query": "{ uspvdb { eia_id ylat xlong } }"}
    try:
        r = requests.post(USPVDB_URL, json=q, timeout=90)
        r.raise_for_status()
        rows = r.json().get("data", {}).get("uspvdb", []) or []
    except Exception as exc:                                   # noqa: BLE001
        print(f"  ⚠  USPVDB unavailable ({type(exc).__name__}: {exc}) — "
              f"solar keeps EIA-860 'facility' precision")
        return {}
    out: dict[int, tuple[float, float]] = {}
    for rec in rows:
        try:
            eia = int(rec["eia_id"])
            out[eia] = (float(rec["ylat"]), float(rec["xlong"]))
        except (TypeError, ValueError, KeyError):
            continue
    print(f"  USPVDB: {len(out):,} solar facilities with verified centroids")
    return out


def fetch_uswtdb() -> dict[int, tuple[float, float]]:
    """EIA plant code → mean position of that project's turbines."""
    try:
        r = requests.get(USWTDB_URL, timeout=120,
                         params={"select": "eia_id,ylat,xlong"})
        r.raise_for_status()
        rows = r.json()
    except Exception as exc:                                   # noqa: BLE001
        print(f"  ⚠  USWTDB unavailable ({type(exc).__name__}: {exc}) — "
              f"wind keeps EIA-860 'facility' precision")
        return {}
    acc: dict[int, list[tuple[float, float]]] = {}
    for rec in rows:
        try:
            eia = int(rec["eia_id"])
            acc.setdefault(eia, []).append((float(rec["ylat"]), float(rec["xlong"])))
        except (TypeError, ValueError, KeyError):
            continue
    out = {k: (sum(p[0] for p in v) / len(v), sum(p[1] for p in v) / len(v))
           for k, v in acc.items()}
    print(f"  USWTDB: {len(out):,} wind projects "
          f"({sum(len(v) for v in acc.values()):,} turbines)")
    return out


# ── main ────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser()
    p.add_argument("--market", required=True, choices=sorted(MARKETS))
    p.add_argument("--apply", action="store_true", help="write to DB (default: dry run)")
    p.add_argument("--limit", type=int, default=None)
    p.add_argument("--skip-usgs", action="store_true")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    cfg = MARKETS[args.market]
    print(f"=== Geolocating {args.market} settlement points "
          f"({'APPLY' if args.apply else 'DRY RUN'}) ===\n")

    conn = psycopg2.connect(DB_URL)
    cur = conn.cursor()
    cur.execute(f"""
        SELECT {cfg['node_col']} AS node,
               MAX(node_type)                AS node_type,
               AVG(avg_da_price)::float      AS avg_da,
               COUNT(*)                      AS months
        FROM {cfg['stats_table']}
        WHERE node_type = 'resource_node'
        GROUP BY 1 ORDER BY 1
        {f'LIMIT {args.limit}' if args.limit else ''}
    """)
    nodes = {r[0]: NodeLoc(node_name=r[0], node_type=r[1],
                           avg_da_price=r[2], months_available=r[3])
             for r in cur.fetchall()}
    print(f"  {len(nodes):,} resource nodes to locate\n")
    if not nodes:
        sys.exit("No resource nodes found — run the node-stats build first.")

    plants, gens = load_eia860(cfg["state"])

    # ── Phase 1: exact LMP node designation ─────────────────────────────────
    by_lmp = {norm(g["lmp_node"]): g for g in gens if g["lmp_node"]}
    p1 = 0
    for name, nl in nodes.items():
        g = by_lmp.get(norm(name))
        if not g:
            continue
        pl = plants.get(g["plant_code"])
        if not pl:
            continue
        nl.latitude, nl.longitude = pl["lat"], pl["lon"]
        nl.precision, nl.method, nl.source = "facility", "reported", "eia860_lmp"
        nl.confidence = 0.90
        nl.eia_plant_code, nl.eia_plant_name = g["plant_code"], pl["name"]
        nl.technology = g["technology"]
        p1 += 1
    print(f"\n  Phase 1 — exact LMP match:        {p1:,} nodes")

    # ── Phase 4 runs BEFORE the USGS upgrades so that fuzzy-matched nodes can
    #    also be upgraded to imagery-verified coordinates.
    by_plant_norm: dict[str, dict] = {}
    for g in gens:
        by_plant_norm.setdefault(norm(g["plant_name"]), g)

    p4 = 0
    for name, nl in nodes.items():
        if nl.precision != "unknown":
            continue
        pref = node_prefix(name)
        if len(pref) < 4:            # too short to match safely
            continue
        best, best_score = None, 0.0
        exact = by_plant_norm.get(pref)
        if exact:
            best, best_score = exact, 100.0
        elif fuzz:
            for pname, g in by_plant_norm.items():
                s = fuzz.ratio(pref, pname)
                if s > best_score:
                    best, best_score = g, s
        # 88 is deliberately strict. Lower thresholds start matching
        # "MOUNTAINVIEW" to "MOUNTAIN VIEW SOLAR" 300km away, and a wrong
        # coordinate is worse than none for a siting tool.
        if not best or best_score < 88:
            continue
        pl = plants.get(best["plant_code"])
        if not pl:
            continue
        nl.latitude, nl.longitude = pl["lat"], pl["lon"]
        nl.precision, nl.method, nl.source = "facility", "name_match", "eia860_name"
        nl.confidence = round(0.50 + 0.25 * (best_score - 88) / 12, 2)
        nl.match_score = round(best_score, 2)
        nl.eia_plant_code, nl.eia_plant_name = best["plant_code"], pl["name"]
        nl.technology = best["technology"]
        p4 += 1
    print(f"  Phase 4 — fuzzy name match (>=88): {p4:,} nodes"
          f"{'' if fuzz else '  [rapidfuzz missing — exact-normalised only]'}")

    # ── Phases 2 & 3: upgrade to imagery-verified coordinates ───────────────
    up_solar = up_wind = 0
    if not args.skip_usgs:
        print()
        pv, wt = fetch_uspvdb(), fetch_uswtdb()
        for nl in nodes.values():
            if nl.eia_plant_code is None:
                continue
            tech = (nl.technology or "").lower()
            if "solar" in tech or "photovoltaic" in tech:
                hit = pv.get(nl.eia_plant_code)
                if hit and nl.would_upgrade_to("exact"):
                    nl.latitude, nl.longitude = hit
                    nl.precision, nl.method, nl.source = "exact", "imagery_verified", "uspvdb"
                    nl.confidence = 0.98
                    up_solar += 1
            elif "wind" in tech:
                hit = wt.get(nl.eia_plant_code)
                if hit and nl.would_upgrade_to("exact"):
                    nl.latitude, nl.longitude = hit
                    nl.precision, nl.method, nl.source = "exact", "imagery_verified", "uswtdb"
                    nl.confidence = 0.98
                    up_wind += 1
        print(f"\n  Phase 2 — USPVDB solar upgrade:   {up_solar:,} nodes")
        print(f"  Phase 3 — USWTDB wind upgrade:    {up_wind:,} nodes")

    # ── Summary ─────────────────────────────────────────────────────────────
    located = [n for n in nodes.values() if n.latitude is not None]
    prec = Counter(n.precision for n in located)
    print(f"\n  ── Result ──")
    print(f"  located:   {len(located):,} / {len(nodes):,} "
          f"({100*len(located)/len(nodes):.1f}%)")
    print(f"  unlocated: {len(nodes)-len(located):,} — left NULL, NOT "
          f"back-filled with county centroids")
    for p, c in sorted(prec.items(), key=lambda kv: -NodeLoc.RANK[kv[0]]):
        print(f"     {p:<10} {c:,}")

    if args.market == "CAISO":
        zones = Counter(caiso_zone_from_latlon(n.latitude, n.longitude) for n in located)
        dlaps = Counter(nearest_dlap(n.latitude, n.longitude) for n in located)
        print(f"  caiso_zone (approx): {dict(zones)}")
        print(f"  dlap (nearest):      {dict(dlaps)}")

    if not args.apply:
        print("\n  DRY RUN — nothing written. Re-run with --apply.")
        for n in located[:8]:
            print(f"     {n.node_name:<28} {n.latitude:9.5f},{n.longitude:10.5f}  "
                  f"{n.precision:<9} {n.source or '':<14} {n.eia_plant_name or ''}")
        cur.close(); conn.close()
        return

    payload = [(
        n.node_name, n.node_type,
        caiso_zone_from_latlon(n.latitude, n.longitude) if args.market == "CAISO" else None,
        nearest_dlap(n.latitude, n.longitude) if args.market == "CAISO" else None,
        n.latitude, n.longitude,
        n.precision, n.method, n.source, n.confidence, n.match_score,
        n.eia_plant_code, n.eia_plant_name, n.technology,
        n.avg_da_price, n.months_available,
    ) for n in located]

    psycopg2.extras.execute_values(cur, f"""
        INSERT INTO {cfg['loc_table']}
          (node_name, node_type, caiso_zone, dlap, latitude, longitude,
           location_precision, location_method, location_source,
           location_confidence, match_score, eia_plant_code, eia_plant_name,
           technology, avg_da_price, months_available)
        VALUES %s
        ON CONFLICT (node_name) DO UPDATE SET
          node_type = EXCLUDED.node_type, caiso_zone = EXCLUDED.caiso_zone,
          dlap = EXCLUDED.dlap, latitude = EXCLUDED.latitude,
          longitude = EXCLUDED.longitude,
          location_precision = EXCLUDED.location_precision,
          location_method = EXCLUDED.location_method,
          location_source = EXCLUDED.location_source,
          location_confidence = EXCLUDED.location_confidence,
          match_score = EXCLUDED.match_score,
          eia_plant_code = EXCLUDED.eia_plant_code,
          eia_plant_name = EXCLUDED.eia_plant_name,
          technology = EXCLUDED.technology,
          avg_da_price = EXCLUDED.avg_da_price,
          months_available = EXCLUDED.months_available,
          updated_at = now()
    """, payload, template="(" + ",".join(["%s"] * 16) + ")", page_size=1000)
    conn.commit()
    print(f"\n  ✓ wrote {len(payload):,} rows to {cfg['loc_table']}")
    cur.close(); conn.close()


if __name__ == "__main__":
    main()
