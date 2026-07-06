#!/usr/bin/env python3
"""
ERCOT Resource Node Geolocation — v2 (4-phase pipeline)

Phase 1 — EIA-860 direct LMP node match
  3_1_Generator_Y2024.xlsx column "RTO/ISO LMP Node Designation" is the ERCOT
  settlement point name. Match it exactly (after stripping dots/spaces) against
  our node_name column. Get lat/lon from 2___Plant_Y2024.xlsx via Plant Code.

Phase 2 — EIA-860 fuzzy plant name match (TX plants)
  For remaining unmatched nodes, fuzzy-match cleaned node prefix against TX plant
  names from the 2024 Plant file (which has more plants than our candidates table).

Phase 3 — ERCOT queue project name fuzzy match
  queue_projects (market=ERCOT) with lat/lon. Match project_name against node_name.

Phase 4 — County centroid fallback
  queue_projects (market=ERCOT) with county but no lat/lon. Fuzzy match project_name
  to node_name, assign Texas county centroid.

Usage:
  cd artifacts/pypsa-engine
  .venv/bin/python3 ../../scripts/src/geo-locate-ercot-nodes-v2.py          # dry run
  .venv/bin/python3 ../../scripts/src/geo-locate-ercot-nodes-v2.py --apply  # write to DB
"""
import io
import os
import re
import sys
import zipfile

import openpyxl
import psycopg2
import psycopg2.extras
from rapidfuzz import fuzz

DATABASE_URL = os.environ["DATABASE_URL"]
EIA860_ZIP   = os.path.join(os.path.dirname(__file__),
                             "../../attached_assets/eia8602024_1777780153233.zip")

# ── Texas county centroids (lat, lon) ─────────────────────────────────────────
TX_COUNTY = {
    "Anderson": (31.82, -95.67), "Andrews": (32.30, -102.64), "Angelina": (31.25, -94.61),
    "Aransas": (28.10, -97.04), "Archer": (33.62, -98.68), "Armstrong": (34.96, -101.36),
    "Atascosa": (28.89, -98.53), "Austin": (29.89, -96.27), "Bailey": (34.07, -102.83),
    "Bandera": (29.75, -99.25), "Bastrop": (30.10, -97.31), "Baylor": (33.61, -99.21),
    "Bee": (28.42, -97.74), "Bell": (31.04, -97.48), "Bexar": (29.45, -98.52),
    "Blanco": (30.26, -98.42), "Borden": (32.74, -101.43), "Bosque": (31.90, -97.63),
    "Bowie": (33.44, -94.17), "Brazoria": (29.17, -95.44), "Brazos": (30.66, -96.30),
    "Brewster": (29.80, -103.25), "Briscoe": (34.53, -100.88), "Brooks": (27.03, -98.22),
    "Brown": (31.77, -99.00), "Burleson": (30.49, -96.65), "Burnet": (30.79, -98.23),
    "Caldwell": (29.84, -97.61), "Calhoun": (28.44, -96.58), "Callahan": (32.30, -99.37),
    "Cameron": (26.15, -97.58), "Camp": (32.97, -94.98), "Carson": (35.40, -101.35),
    "Cass": (33.08, -94.34), "Castro": (34.53, -102.27), "Chambers": (29.71, -94.74),
    "Cherokee": (31.84, -95.16), "Childress": (34.53, -100.21), "Clay": (33.78, -98.21),
    "Cochran": (33.66, -102.83), "Coke": (31.89, -100.53), "Coleman": (31.82, -99.43),
    "Collin": (33.19, -96.57), "Collingsworth": (34.96, -100.27), "Colorado": (29.62, -96.52),
    "Comal": (29.80, -98.26), "Comanche": (31.95, -98.56), "Concho": (31.32, -99.88),
    "Cooke": (33.64, -97.21), "Corpus Christi": (27.80, -97.40),
    "Coryell": (31.39, -97.79), "Cottle": (34.08, -100.28), "Crane": (31.43, -102.35),
    "Crockett": (30.72, -101.42), "Crosby": (33.66, -101.30), "Culberson": (30.84, -104.52),
    "Dallam": (36.28, -102.60), "Dallas": (32.77, -96.79), "Dawson": (32.74, -101.95),
    "Deaf Smith": (34.96, -102.60), "Delta": (33.39, -95.67), "Denton": (33.21, -97.12),
    "DeWitt": (29.09, -97.35), "Dickens": (33.66, -100.79), "Dimmit": (28.43, -99.75),
    "Donley": (34.96, -100.82), "Duval": (27.68, -98.52), "Eastland": (32.30, -98.83),
    "Ector": (31.87, -102.53), "Edwards": (29.97, -100.30), "Ellis": (32.35, -96.80),
    "El Paso": (31.77, -106.27), "Erath": (32.23, -98.22), "Falls": (31.26, -96.93),
    "Fannin": (33.59, -96.11), "Fayette": (29.88, -96.92), "Fisher": (32.74, -100.40),
    "Floyd": (34.07, -101.31), "Foard": (33.99, -99.78), "Fort Bend": (29.53, -95.77),
    "Franklin": (33.17, -95.22), "Freestone": (31.70, -96.15), "Frio": (28.87, -99.11),
    "Gaines": (32.74, -102.64), "Galveston": (29.28, -94.80), "Garza": (33.18, -101.30),
    "Gillespie": (30.32, -98.95), "Glasscock": (31.87, -101.52), "Goliad": (28.65, -97.38),
    "Gonzales": (29.45, -97.49), "Gray": (35.40, -100.82), "Grayson": (33.62, -96.68),
    "Gregg": (32.48, -94.82), "Grimes": (30.55, -95.98), "Guadalupe": (29.59, -97.95),
    "Hale": (34.07, -101.82), "Hall": (34.53, -100.68), "Hamilton": (31.70, -98.11),
    "Hansford": (36.28, -101.35), "Hardeman": (34.29, -99.75), "Hardin": (30.34, -94.38),
    "Harris": (29.85, -95.40), "Harrison": (32.55, -94.37), "Hartley": (35.84, -102.60),
    "Haskell": (33.16, -99.73), "Hays": (30.06, -98.03), "Hemphill": (35.84, -100.27),
    "Henderson": (32.21, -95.85), "Hidalgo": (26.30, -98.18), "Hill": (31.99, -97.13),
    "Hockley": (33.61, -102.34), "Hood": (32.43, -97.83), "Hopkins": (33.15, -95.56),
    "Houston": (31.32, -95.43), "Howard": (32.30, -101.44), "Hudspeth": (31.46, -105.40),
    "Hunt": (33.13, -96.08), "Hutchinson": (35.84, -101.35), "Irion": (31.29, -100.98),
    "Jack": (33.23, -98.17), "Jackson": (28.96, -96.58), "Jasper": (30.75, -94.01),
    "Jeff Davis": (30.72, -104.13), "Jefferson": (29.84, -94.16), "Jim Hogg": (27.03, -98.70),
    "Jim Wells": (27.73, -98.09), "Johnson": (32.38, -97.37), "Jones": (32.74, -99.88),
    "Karnes": (28.90, -97.86), "Kaufman": (32.60, -96.28), "Kendall": (29.94, -98.71),
    "Kenedy": (26.93, -97.62), "Kent": (33.18, -100.79), "Kerr": (30.06, -99.34),
    "Kimble": (30.49, -99.75), "King": (33.62, -100.26), "Kinney": (29.35, -100.42),
    "Kleberg": (27.43, -97.68), "Knox": (33.61, -99.75), "Lamar": (33.67, -95.57),
    "Lamb": (34.07, -102.35), "Lampasas": (31.20, -98.23), "La Salle": (28.34, -99.10),
    "Lavaca": (29.38, -96.93), "Lee": (30.31, -96.97), "Leon": (31.30, -95.99),
    "Liberty": (30.16, -94.81), "Limestone": (31.55, -96.58), "Lipscomb": (36.28, -100.27),
    "Live Oak": (28.35, -98.13), "Llano": (30.71, -98.67), "Loving": (31.85, -103.60),
    "Lubbock": (33.61, -101.82), "Lynn": (33.18, -101.82), "Madison": (30.96, -95.92),
    "Marion": (32.79, -94.36), "Martin": (32.30, -101.95), "Mason": (30.71, -99.23),
    "Matagorda": (28.78, -95.99), "Maverick": (28.74, -100.30), "McCulloch": (31.20, -99.35),
    "McLennan": (31.55, -97.19), "McMullen": (28.35, -98.57), "Medina": (29.35, -99.11),
    "Menard": (30.88, -99.82), "Midland": (31.87, -102.03), "Milam": (30.79, -96.97),
    "Mills": (31.49, -98.60), "Mitchell": (32.30, -100.92), "Montague": (33.68, -97.72),
    "Montgomery": (30.30, -95.50), "Moore": (35.84, -101.89), "Morris": (33.12, -94.73),
    "Motley": (34.07, -100.79), "Nacogdoches": (31.61, -94.63), "Navarro": (32.04, -96.47),
    "Newton": (30.78, -93.74), "Nolan": (32.30, -100.40), "Nueces": (27.73, -97.55),
    "Ochiltree": (36.28, -100.82), "Oldham": (35.40, -102.60), "Orange": (30.12, -93.90),
    "Palo Pinto": (32.75, -98.31), "Panola": (32.16, -94.31), "Parker": (32.78, -97.80),
    "Parmer": (34.53, -102.78), "Pecos": (30.79, -102.72), "Polk": (30.79, -94.83),
    "Potter": (35.40, -101.89), "Presidio": (29.73, -104.13), "Rains": (32.87, -95.80),
    "Randall": (34.96, -101.89), "Reagan": (31.37, -101.52), "Real": (29.83, -99.82),
    "Red River": (33.62, -94.99), "Reeves": (31.32, -103.69), "Refugio": (28.31, -97.16),
    "Roberts": (35.84, -100.82), "Robertson": (31.03, -96.52), "Rockwall": (32.90, -96.46),
    "Runnels": (31.83, -99.97), "Rusk": (32.11, -94.77), "Sabine": (31.34, -93.85),
    "San Augustine": (31.39, -94.17), "San Jacinto": (30.58, -95.16),
    "San Patricio": (27.98, -97.52), "San Saba": (31.15, -98.72), "Schleicher": (30.90, -100.54),
    "Scurry": (32.74, -100.92), "Shackelford": (32.74, -99.37), "Shelby": (31.80, -94.15),
    "Sherman": (36.28, -101.89), "Smith": (32.37, -95.27), "Somervell": (32.22, -97.77),
    "Starr": (26.56, -98.74), "Stephens": (32.74, -98.82), "Sterling": (31.83, -100.99),
    "Stonewall": (33.18, -100.26), "Sutton": (30.49, -100.54), "Swisher": (34.53, -101.73),
    "Tarrant": (32.77, -97.29), "Taylor": (32.30, -99.88), "Terrell": (30.22, -102.07),
    "Terry": (33.18, -102.34), "Throckmorton": (33.18, -99.21), "Titus": (33.21, -94.97),
    "Tom Green": (31.41, -100.46), "Travis": (30.33, -97.77), "Trinity": (31.09, -95.13),
    "Tyler": (30.78, -94.38), "Upshur": (32.73, -94.97), "Upton": (31.37, -102.03),
    "Uvalde": (29.35, -99.75), "Val Verde": (29.89, -101.16), "Van Zandt": (32.56, -95.84),
    "Victoria": (28.79, -96.98), "Walker": (30.74, -95.57), "Waller": (29.89, -95.99),
    "Ward": (31.51, -103.11), "Washington": (30.22, -96.40), "Webb": (27.76, -99.49),
    "Wharton": (29.27, -96.21), "Wheeler": (35.40, -100.27), "Wichita": (33.90, -98.69),
    "Wilbarger": (34.08, -99.21), "Willacy": (26.48, -97.83), "Williamson": (30.65, -97.60),
    "Wilson": (29.17, -98.09), "Winkler": (31.83, -103.06), "Wise": (33.22, -97.66),
    "Wood": (32.78, -95.37), "Yoakum": (33.18, -102.83), "Young": (33.18, -98.68),
    "Zapata": (26.99, -99.17), "Zavala": (28.87, -99.75),
}


# ── Normalization helpers ──────────────────────────────────────────────────────

TECH_TOKENS_NODE = re.compile(
    r'\b(SLR|WND|WIND|ESS|BESS|RN|ALL|UNIT\d*|DGR\d*|G\d+|ES\d*|_\d+)\b',
    re.IGNORECASE,
)
TECH_TOKENS_EIA = re.compile(
    r'\b(Wind|Solar|Farm|Project|LLC|LP|Energy|Power|Plant|Station|'
    r'Storage|Battery|BESS|ESS|Hybrid|Holdings|Partners|Resource|'
    r'I{1,3}|II|III|IV|V|VI|VII|VIII|IX|X|[0-9]+)\b',
    re.IGNORECASE,
)


def clean_node(node_name: str) -> str:
    prefix = node_name.split('_')[0]
    parts  = node_name.split('_')
    prefix2 = parts[0] + (parts[1] if len(parts) > 1 else '')
    cleaned = TECH_TOKENS_NODE.sub('', prefix).strip()
    return (cleaned or prefix).lower()


def clean_eia(name: str) -> str:
    cleaned = TECH_TOKENS_EIA.sub(' ', name)
    return re.sub(r'\s+', ' ', cleaned).strip().lower()


def fuzzy_score(nc: str, ec: str) -> float:
    if len(nc) < 4:
        return 0.0  # too short for reliable fuzzy match
    if len(ec) < max(4, len(nc) - 2):
        return 0.0
    ts = fuzz.token_sort_ratio(nc, ec)
    r  = fuzz.ratio(nc, ec)
    if len(nc) <= 5:
        # partial_ratio is misleading for short strings (e.g. 'ang' in 'langer')
        # weight token_sort + ratio only
        return ts * 0.5 + r * 0.5
    pr = fuzz.partial_ratio(nc, ec)
    return pr * 0.5 + ts * 0.3 + r * 0.2


# ── Load EIA-860 data ─────────────────────────────────────────────────────────

def load_eia860(zip_path: str):
    """Return (plants_dict, generators_list) for TX plants.

    plants_dict: {plant_code -> (plant_name, lat, lon)}
    generators_list: [(plant_code, plant_name, lmp_node_norm, lat, lon)]
    """
    plants_dict: dict[int, tuple] = {}
    generators:  list[tuple]      = []

    with zipfile.ZipFile(zip_path) as z:
        # ── Plant file ────────────────────────────────────────────────────────
        plant_data = z.read('2___Plant_Y2024.xlsx')
        wb = openpyxl.load_workbook(io.BytesIO(plant_data), read_only=True, data_only=True)
        ws = wb.active
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if headers is None:
                headers = list(row)
                pc_i   = headers.index('Plant Code')
                pn_i   = headers.index('Plant Name')
                st_i   = headers.index('State')
                lat_i  = headers.index('Latitude')
                lon_i  = headers.index('Longitude')
                continue
            if row[st_i] != 'TX':
                continue
            try:
                pc  = int(row[pc_i])
                lat = float(row[lat_i])
                lon = float(row[lon_i])
                plants_dict[pc] = (row[pn_i], lat, lon)
            except (TypeError, ValueError):
                pass
        wb.close()

        # ── Generator file ────────────────────────────────────────────────────
        gen_data = z.read('3_1_Generator_Y2024.xlsx')
        wb = openpyxl.load_workbook(io.BytesIO(gen_data), read_only=True, data_only=True)
        ws = wb.active
        headers = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if headers is None:
                headers = list(row)
                pc_i   = headers.index('Plant Code')
                pn_i   = headers.index('Plant Name')
                st_i   = headers.index('State')
                lmp_i  = headers.index('RTO/ISO LMP Node Designation')
                continue
            if row[st_i] != 'TX':
                continue
            lmp = row[lmp_i]
            if not lmp:
                continue
            try:
                pc = int(row[pc_i])
            except (TypeError, ValueError):
                continue
            if pc not in plants_dict:
                continue
            pname, lat, lon = plants_dict[pc]
            # Normalize the LMP designation: uppercase, strip dots/spaces/dashes
            lmp_norm = re.sub(r'[\s.\-]', '', str(lmp)).upper()
            generators.append((pc, pname, lmp_norm, lat, lon))
        wb.close()

    print(f"EIA-860: {len(plants_dict)} TX plants, {len(generators)} TX generators with LMP nodes")
    return plants_dict, generators


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    dry_run = '--apply' not in sys.argv

    conn = psycopg2.connect(DATABASE_URL)
    cur  = conn.cursor()

    # Load unmatched nodes
    cur.execute("""
        SELECT id, node_name, load_zone
        FROM ercot_node_locations
        WHERE node_type = 'resource_node' AND location_source = 'zone_centroid'
        ORDER BY node_name
    """)
    unmatched = cur.fetchall()
    print(f"Unmatched nodes: {len(unmatched)}")

    # Build a dict for O(1) lookup: normalised_key → [(node_id, node_name, load_zone)]
    # Normalise: uppercase, strip ., spaces, dashes
    def norm_key(s: str) -> str:
        return re.sub(r'[\s.\-_]', '', s).upper()

    node_by_norm: dict[str, list] = {}
    for nid, nname, lz in unmatched:
        k = norm_key(nname)
        node_by_norm.setdefault(k, []).append((nid, nname, lz))

    remaining = list(unmatched)  # shrinks as we match
    matched_ids: set[int] = set()
    updates: list[dict] = []

    # ─── Phase 1: EIA-860 direct LMP match ────────────────────────────────────
    print("\n── Phase 1: EIA-860 direct LMP node match ──────────────────────────")
    plants_dict, generators = load_eia860(EIA860_ZIP)

    # Build a generator lookup: lmp_norm → (plant_name, lat, lon)
    gen_lookup: dict[str, tuple] = {}
    for pc, pname, lmp_norm, lat, lon in generators:
        gen_lookup[lmp_norm] = (pname, lat, lon)

    phase1_hits = 0
    for nid, nname, lz in list(remaining):
        k = norm_key(nname)
        if k in gen_lookup:
            pname, lat, lon = gen_lookup[k]
            updates.append({
                'id': nid, 'lat': lat, 'lon': lon,
                'source': 'eia_lmp_direct',
                'eia_plant_name': pname,
                'phase': 1,
            })
            matched_ids.add(nid)
            phase1_hits += 1
            print(f"  {nname:30s} → {pname:40s} ({lat:.4f},{lon:.4f})")

    remaining = [(nid, nn, lz) for nid, nn, lz in remaining if nid not in matched_ids]
    print(f"Phase 1: {phase1_hits} direct matches  |  {len(remaining)} still unmatched")

    # ─── Phase 2: EIA-860 fuzzy plant name match ──────────────────────────────
    print("\n── Phase 2: EIA-860 fuzzy plant name match ─────────────────────────")
    FUZZY_THRESH = 80

    # Build EIA plant records for fuzzy matching
    plant_records = [(pname, lat, lon, clean_eia(pname))
                     for pname, lat, lon in plants_dict.values()
                     if lat and lon]

    phase2_hits = 0
    for nid, nname, lz in list(remaining):
        nc = clean_node(nname)
        if len(nc) < 3:
            continue
        best_s, best_r = 0.0, None
        for pname, lat, lon, pc in plant_records:
            s = fuzzy_score(nc, pc)
            if s > best_s:
                best_s, best_r = s, (pname, lat, lon)
        if best_r and best_s >= FUZZY_THRESH:
            updates.append({
                'id': nid, 'lat': best_r[1], 'lon': best_r[2],
                'source': 'eia_fuzzy_match',
                'eia_plant_name': best_r[0],
                'phase': 2,
            })
            matched_ids.add(nid)
            phase2_hits += 1
            if phase2_hits <= 30:
                print(f"  {nname:30s} → {best_r[0]:40s}  score={best_s:.0f}")

    remaining = [(nid, nn, lz) for nid, nn, lz in remaining if nid not in matched_ids]
    print(f"Phase 2: {phase2_hits} fuzzy matches  |  {len(remaining)} still unmatched")

    # ─── Phase 3: ERCOT queue lat/lon match ───────────────────────────────────
    print("\n── Phase 3: ERCOT queue project lat/lon match ──────────────────────")
    cur.execute("""
        SELECT project_name, latitude, longitude, county
        FROM queue_projects
        WHERE market = 'ERCOT'
          AND latitude IS NOT NULL AND latitude != 0
          AND longitude IS NOT NULL AND longitude != 0
    """)
    queue_with_latlon = [(r[0], float(r[1]), float(r[2]), r[3]) for r in cur.fetchall()]

    queue_records = [(pn, lat, lon, county, clean_eia(pn))
                     for pn, lat, lon, county in queue_with_latlon]

    QUEUE_THRESH = 75  # slightly lower — queue names are noisier
    phase3_hits = 0
    for nid, nname, lz in list(remaining):
        nc = clean_node(nname)
        if len(nc) < 3:
            continue
        best_s, best_r = 0.0, None
        for pn, lat, lon, county, pc in queue_records:
            s = fuzzy_score(nc, pc)
            if s > best_s:
                best_s, best_r = s, (pn, lat, lon)
        if best_r and best_s >= QUEUE_THRESH:
            updates.append({
                'id': nid, 'lat': best_r[1], 'lon': best_r[2],
                'source': 'queue_latlon_match',
                'eia_plant_name': best_r[0],
                'phase': 3,
            })
            matched_ids.add(nid)
            phase3_hits += 1
            if phase3_hits <= 20:
                print(f"  {nname:30s} → {best_r[0]:40s}  score={best_s:.0f}")

    remaining = [(nid, nn, lz) for nid, nn, lz in remaining if nid not in matched_ids]
    print(f"Phase 3: {phase3_hits} queue lat/lon matches  |  {len(remaining)} still unmatched")

    # ─── Phase 4: County centroid fallback ────────────────────────────────────
    print("\n── Phase 4: County centroid fallback ───────────────────────────────")
    cur.execute("""
        SELECT project_name, county
        FROM queue_projects
        WHERE market = 'ERCOT' AND county IS NOT NULL
    """)
    queue_county = [(r[0], r[1]) for r in cur.fetchall()]
    county_records = [(pn, county, clean_eia(pn)) for pn, county in queue_county if county]

    COUNTY_THRESH = 72
    phase4_hits = 0
    for nid, nname, lz in list(remaining):
        nc = clean_node(nname)
        if len(nc) < 3:
            continue
        best_s, best_county, best_pname = 0.0, None, None
        for pn, county, pc in county_records:
            s = fuzzy_score(nc, pc)
            if s > best_s and county in TX_COUNTY:
                best_s, best_county, best_pname = s, county, pn
        if best_county and best_s >= COUNTY_THRESH:
            clat, clon = TX_COUNTY[best_county]
            updates.append({
                'id': nid, 'lat': clat, 'lon': clon,
                'source': 'county_centroid',
                'eia_plant_name': f"{best_pname} (county: {best_county})",
                'phase': 4,
            })
            matched_ids.add(nid)
            phase4_hits += 1
            if phase4_hits <= 20:
                print(f"  {nname:30s} → county={best_county:15s}  score={best_s:.0f}")

    remaining = [(nid, nn, lz) for nid, nn, lz in remaining if nid not in matched_ids]
    print(f"Phase 4: {phase4_hits} county centroid matches  |  {len(remaining)} still unmatched")

    # ─── Summary ──────────────────────────────────────────────────────────────
    print(f"\n{'='*60}")
    print(f"Total new matches: {len(updates)}")
    by_phase = {}
    for u in updates:
        by_phase[u['phase']] = by_phase.get(u['phase'], 0) + 1
    for p, n in sorted(by_phase.items()):
        src = {1:'eia_lmp_direct', 2:'eia_fuzzy_match', 3:'queue_latlon_match', 4:'county_centroid'}[p]
        print(f"  Phase {p} ({src}): {n}")
    print(f"Still unmatched after all phases: {len(remaining)}")

    if dry_run:
        print("\n[DRY RUN] Pass --apply to write to DB")
        cur.close(); conn.close(); return

    # ─── Apply ────────────────────────────────────────────────────────────────
    print(f"\nApplying {len(updates)} updates...")
    for u in updates:
        cur.execute("""
            UPDATE ercot_node_locations
            SET latitude        = %s,
                longitude       = %s,
                location_source = %s,
                eia_plant_name  = %s
            WHERE id = %s
        """, (u['lat'], u['lon'], u['source'], u['eia_plant_name'], u['id']))
    conn.commit()
    print("Done.")

    cur.execute("""
        SELECT location_source, COUNT(*)
        FROM ercot_node_locations WHERE node_type = 'resource_node'
        GROUP BY location_source ORDER BY count DESC
    """)
    print("\nFinal location breakdown:")
    total_geo = 0
    for row in cur.fetchall():
        print(f"  {row[0]:30s}: {row[1]}")
        if row[0] != 'zone_centroid':
            total_geo += row[1]
    print(f"\nTotal geo-located (non-centroid): {total_geo}")

    cur.close()
    conn.close()


if __name__ == '__main__':
    main()
