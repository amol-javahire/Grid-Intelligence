#!/usr/bin/env python3
"""
Seed AESO transmission capability data.

AUTHORITATIVE SOURCE — the WORKBOOK, not the PDF
---------------------------------------------------------------------------
  Transmission-Capability-Results-Sept-2025.xlsx
    sheet "Substation Capabilities Table"  213 substations, wide format
    sheet "Line Capabilities Table"        491 lines

AESO publishes the same results as a PDF report and as this workbook. The
workbook is structured and exact; the PDF requires coordinate-based scraping
that produced real errors — an early parse read Anderson's second bus as
138 kV when the workbook says 144 kV, glued "549" and "041" into one bus
number, and bled planning-area text between adjacent rows. Do NOT parse the
PDF for A/B data. infra/parse-aeso-capability-report.py is retained only for
Attachment C (2024-energized assets), which has no workbook equivalent.

  Download page:
  https://www.aeso.ca/grid/connecting-to-the-grid/transmission-capability-map/

STUDY AREA — CRITICAL SCOPE LIMIT (report section 2.3)
  This assessment covers ONLY the South and Central East planning regions:
  Medicine Hat (4), Lloydminster (13), Wainwright (32), Alliance/Battle River
  (36), Provost (37), Hanna (42), Sheerness (43), Strathmore/Blackie (45),
  High River (46), Brooks (47), Empress (48), Stavely (49), Vauxhall (52),
  Fort Macleod (53), Lethbridge (54), Glenwood (55), Vegreville (56) — plus a
  single Calgary (6) bus. AESO chose these because they hold most of the
  active connection-project interest.

  There is NO capability data for Edmonton, Wabamun, Fort McMurray, Grande
  Prairie, Peace River or anywhere else north/west. So this CANNOT by itself
  supply line ratings for the whole 9-bus provincial model — it covers roughly
  the South region and the eastern half of Central. Any UI must scope its
  claims to the study area, and the PyPSA model must keep another source for
  the northern corridors.

WHAT THE NUMBERS MEAN
  Substation Capability (MW): additional generation connectable at that bus
  before N-0 (category A) thermal congestion, at the 0.5 percentile of the
  historical duration curve. AESO states these are INDICATIVE and not
  guaranteed in the Connection Process. They are a screening signal, not an
  entitlement — label them that way in any UI.

  Line Capability (MW): same concept applied to each line, reported per
  connecting substation. A line appears once per endpoint, so 491 rows cover
  fewer than 491 physical lines.

SHAPE
  The substation sheet is WIDE: one row per substation with up to four
  (Bus Number, Voltage kV, Capability MW) triples. This seeder normalises it
  to one row per (substation, bus) — 213 substations expand to ~236 buses.

Requires: openpyxl, psycopg2
Usage:
  python infra/seed-aeso-capability.py --inspect     # read-only
  python infra/seed-aeso-capability.py --write
"""

import os
import re
import sys
import logging
from typing import Optional

try:
    import openpyxl
except ImportError:
    print("openpyxl missing. Run: artifacts/pypsa-engine/.venv/bin/pip install openpyxl")
    sys.exit(1)

logging.basicConfig(level=logging.INFO, format="%(message)s")
log = logging.getLogger("aeso-capability")

XLSX = os.environ.get("AESO_CAPABILITY_XLSX", "Transmission-Capability-Results-Sept-2025.xlsx")
SOURCE_DOC = "AESO Transmission Capability Results, Sept 2025 Assessment (workbook)"
AS_OF = "2025-09-26"

SHEET_SUB = "Substation Capabilities Table"
SHEET_LINE = "Line Capabilities Table"


def _int(v) -> Optional[int]:
    if v is None or str(v).strip() == "":
        return None
    try:
        return int(float(str(v).strip()))
    except ValueError:
        return None


def _num(v) -> Optional[float]:
    if v is None or str(v).strip() == "":
        return None
    try:
        return float(str(v).strip())
    except ValueError:
        return None


def _area(v) -> tuple[Optional[int], Optional[str]]:
    """
    '04-Medicine Hat' -> (4, 'Medicine Hat').
    The numeric code is the authoritative key — names vary in punctuation
    ('36-Alliance/Battle River') and are display-only.
    """
    if not v:
        return None, None
    s = str(v).strip()
    m = re.match(r"^(\d{1,2})\s*-\s*(.+)$", s)
    if m:
        return int(m.group(1)), m.group(2).strip()
    return None, s


def read_workbook():
    if not os.path.exists(XLSX):
        log.error("Workbook not found: %s (set AESO_CAPABILITY_XLSX)", XLSX)
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX, data_only=True)

    # ── Substations: wide (4 bus blocks) → long (one row per bus) ──────────
    ws = wb[SHEET_SUB]
    subs = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        name, code, tfo, area = row[0], row[1], row[2], row[3]
        if not code:
            continue
        area_code, area_name = _area(area)
        for block in range(4):
            base = 4 + block * 3
            bus, kv, cap = _int(row[base]), _int(row[base + 1]), _num(row[base + 2])
            if bus is None:
                continue
            subs.append({
                "facility_name": str(name).strip() if name else None,
                "facility_code": str(code).strip(),
                "tfo": str(tfo).strip() if tfo else None,
                "planning_area_code": area_code,
                "planning_area_name": area_name,
                "bus_number": bus,
                "voltage_kv": kv,
                "capability_mw": cap,
            })

    # ── Lines ──────────────────────────────────────────────────────────────
    ws = wb[SHEET_LINE]
    lines = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        lname = row[0]
        if not lname:
            continue
        area_code, area_name = _area(row[4])
        # "1002L [275S-163S]" → endpoints 275S, 163S. This is the network
        # topology: facility code → facility code, which is what lets the
        # 9-bus model use real corridor ratings instead of guesses.
        m = re.search(r"\[([\w\-]+?)-([\w\-]+?)\]", str(lname))
        lines.append({
            "line_name": str(lname).strip(),
            "voltage_kv": _int(row[1]),
            "substation_name": str(row[2]).strip() if row[2] else None,
            "facility_code": str(row[3]).strip() if row[3] else None,
            "planning_area_code": area_code,
            "planning_area_name": area_name,
            "tfo": str(row[5]).strip() if row[5] else None,
            "capability_mw": _num(row[6]),
            "endpoint_a": m.group(1) if m else None,
            "endpoint_b": m.group(2) if m else None,
        })

    return subs, lines


def inspect(subs, lines):
    log.info("=== Substations: %d buses across %d facilities ===",
             len(subs), len({s["facility_code"] for s in subs}))
    for s in subs[:10]:
        log.info("  %-24s %-6s %-9s area=%-3s %-20s bus=%-7s %4skV  %6s MW",
                 (s["facility_name"] or "")[:24], s["facility_code"], (s["tfo"] or "")[:9],
                 s["planning_area_code"], (s["planning_area_name"] or "")[:20],
                 s["bus_number"], s["voltage_kv"], s["capability_mw"])

    log.info("\n=== Lines: %d rows, %d distinct line names ===",
             len(lines), len({l["line_name"] for l in lines}))
    for l in lines[:10]:
        log.info("  %-24s %4skV %-20s %-6s area=%-3s %6s MW  [%s→%s]",
                 l["line_name"][:24], l["voltage_kv"], (l["substation_name"] or "")[:20],
                 l["facility_code"], l["planning_area_code"], l["capability_mw"],
                 l["endpoint_a"], l["endpoint_b"])

    areas = {}
    for s in subs:
        if s["planning_area_code"] is not None:
            areas[s["planning_area_code"]] = s["planning_area_name"]
    log.info("\n=== Planning areas: %d ===", len(areas))
    for k in sorted(areas):
        n = len([s for s in subs if s["planning_area_code"] == k])
        mw = sum(s["capability_mw"] or 0 for s in subs if s["planning_area_code"] == k)
        log.info("    %02d-%-24s %3d buses  %8.0f MW total headroom", k, areas[k], n, mw)

    log.info("\n=== Quality ===")
    log.info("  buses missing capability : %d", len([s for s in subs if s["capability_mw"] is None]))
    log.info("  buses missing voltage    : %d", len([s for s in subs if s["voltage_kv"] is None]))
    log.info("  lines missing capability : %d", len([l for l in lines if l["capability_mw"] is None]))
    log.info("  lines w/o parsed endpoints: %d", len([l for l in lines if not l["endpoint_a"]]))
    log.info("  distinct bus numbers     : %d (of %d rows)",
             len({s["bus_number"] for s in subs}), len(subs))
    kv = sorted({s["voltage_kv"] for s in subs if s["voltage_kv"]})
    log.info("  voltage levels           : %s", kv)
    caps = [s["capability_mw"] for s in subs if s["capability_mw"] is not None]
    log.info("  headroom MW              : min %.0f  max %.0f  total %.0f",
             min(caps), max(caps), sum(caps))
    zero = len([c for c in caps if c == 0])
    log.info("  buses with ZERO headroom : %d (%.0f%%) — fully congested for new gen",
             zero, 100 * zero / len(caps))


def write(subs, lines):
    import psycopg2, psycopg2.extras
    dsn = os.environ.get("DATABASE_URL")
    if not dsn:
        log.error("DATABASE_URL not set")
        sys.exit(1)
    conn = psycopg2.connect(dsn)
    with conn.cursor() as cur:
        cur.execute("""
        CREATE TABLE IF NOT EXISTS aeso_substation_capability (
            id serial PRIMARY KEY,
            facility_name text, facility_code text, tfo text,
            planning_area_code int, planning_area_name text,
            bus_number int, voltage_kv int, capability_mw numeric(10,2),
            source_document text, as_of_date date,
            UNIQUE (facility_code, bus_number, as_of_date));
        CREATE INDEX IF NOT EXISTS aeso_subcap_area_idx  ON aeso_substation_capability (planning_area_code);
        CREATE INDEX IF NOT EXISTS aeso_subcap_code_idx  ON aeso_substation_capability (facility_code);

        CREATE TABLE IF NOT EXISTS aeso_line_capability (
            id serial PRIMARY KEY,
            line_name text, voltage_kv int, substation_name text, facility_code text,
            planning_area_code int, planning_area_name text, tfo text,
            capability_mw numeric(10,2), endpoint_a text, endpoint_b text,
            source_document text, as_of_date date,
            UNIQUE (line_name, facility_code, as_of_date));
        CREATE INDEX IF NOT EXISTS aeso_linecap_area_idx ON aeso_line_capability (planning_area_code);
        CREATE INDEX IF NOT EXISTS aeso_linecap_ep_idx   ON aeso_line_capability (endpoint_a, endpoint_b);
        """)

        psycopg2.extras.execute_batch(cur, """
            INSERT INTO aeso_substation_capability
              (facility_name, facility_code, tfo, planning_area_code, planning_area_name,
               bus_number, voltage_kv, capability_mw, source_document, as_of_date)
            VALUES (%(facility_name)s,%(facility_code)s,%(tfo)s,%(planning_area_code)s,
                    %(planning_area_name)s,%(bus_number)s,%(voltage_kv)s,%(capability_mw)s,
                    %(src)s,%(asof)s)
            ON CONFLICT (facility_code, bus_number, as_of_date) DO UPDATE SET
              capability_mw = EXCLUDED.capability_mw,
              voltage_kv    = EXCLUDED.voltage_kv
        """, [{**r, "src": SOURCE_DOC, "asof": AS_OF} for r in subs])

        psycopg2.extras.execute_batch(cur, """
            INSERT INTO aeso_line_capability
              (line_name, voltage_kv, substation_name, facility_code, planning_area_code,
               planning_area_name, tfo, capability_mw, endpoint_a, endpoint_b,
               source_document, as_of_date)
            VALUES (%(line_name)s,%(voltage_kv)s,%(substation_name)s,%(facility_code)s,
                    %(planning_area_code)s,%(planning_area_name)s,%(tfo)s,%(capability_mw)s,
                    %(endpoint_a)s,%(endpoint_b)s,%(src)s,%(asof)s)
            ON CONFLICT (line_name, facility_code, as_of_date) DO UPDATE SET
              capability_mw = EXCLUDED.capability_mw
        """, [{**r, "src": SOURCE_DOC, "asof": AS_OF} for r in lines])
    conn.commit()
    conn.close()
    log.info("\nWrote %d substation buses and %d line rows.", len(subs), len(lines))


if __name__ == "__main__":
    s, l = read_workbook()
    inspect(s, l)
    if "--write" in sys.argv:
        write(s, l)
    else:
        log.info("\n(read-only — pass --write to load into the database)")
