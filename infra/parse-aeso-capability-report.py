#!/usr/bin/env python3
"""
Parse the AESO Transmission Capability Map Report PDF into three tables.

SOURCE: Transmission-Capability-Map-Report-Sept-2025.pdf (AESO, 2025 Assessment,
dated 26 Sep 2025, Classification: Public). Download page:
https://www.aeso.ca/grid/connecting-to-the-grid/transmission-capability-map/

WHAT THIS REPLACES
---------------------------------------------------------------------------
The 9-bus AESO PyPSA model (aeso_network_regional.py) previously carried five
line ratings tagged "estimated:" — engineering guesses of 3,000–3,500 MW that
never bind, which is why LMPs were identical at every bus regardless of load.
aeso_generators.py additionally placed generators using a hand-built plant-name
map because AESO's API publishes no location data. Both are superseded by this
document, which is AESO's own published assessment.

THE THREE ATTACHMENTS
  A (p9-18,  ~238 rows) Substation Capability
      Facility Name | Code (163S) | TFO | Planning Area (48-Empress)
      | Bus Number | Voltage kV | Capability MW
      Capability MW = additional generation connectable before N-0 thermal
      congestion at the 0.5 percentile. This is a siting metric in its own
      right, not just a model input.
  B (p19-42, ~487 rows) Transmission Line Capability
      Line Name (1038L [266S-138S]) | Voltage kV | Substation Name
      | Facility Code | Planning Area | TFO | Capability MW
  C (p43,    ~15 rows)  Generation Assets Energized in 2024
      Asset Name (ASSET_ID) | Max Capability Change MW | Area | Region
      The Area->Region rollup here is the ONLY published mapping of AESO
      planning areas to the six planning regions the 9-bus model uses.

PARSING NOTES — why coordinates and not extract_tables()
  pdfplumber's table extractor returns heavily fragmented cells on this
  document. Rows are therefore rebuilt from word coordinates: words are
  clustered into physical lines by their `top`, assigned to columns by `x0`
  against the header positions, and multi-line records merged.

  The Planning Area column is narrow and wraps MID-WORD ("36-" / "Alliance/
  Battl" / "e River"). Joining those with spaces produces garbage, so the
  AREA CODE (the integer before the hyphen) is parsed as the authoritative
  key and the concatenated name kept only for display. Never join on the name.

  --inspect prints parsed rows and counts WITHOUT touching the database.
  Always run that first and spot-check against the PDF.

Requires: pdfplumber  (venv: artifacts/pypsa-engine/.venv)
  artifacts/pypsa-engine/.venv/bin/pip install pdfplumber

Usage:
  python infra/parse-aeso-capability-report.py --inspect
  python infra/parse-aeso-capability-report.py --write
"""

import os
import re
import sys
import logging
from collections import defaultdict
from typing import Optional

try:
    import openpyxl
except ImportError:
    print("openpyxl not installed. Run:\n"
          "  artifacts/pypsa-engine/.venv/bin/pip install openpyxl")
    sys.exit(1)

logging.basicConfig(level=logging.INFO, format="%(message)s")
log = logging.getLogger("aeso-capability")

PDF = os.environ.get("AESO_CAPABILITY_PDF", "Transmission-Capability-Map-Report-Sept-2025.pdf")
WORKBOOK = os.environ.get(
    "AESO_CAPABILITY_XLSX", "Transmission-Capability-Results-Sept-2025.xlsx"
)
SOURCE_DOC = "AESO Transmission Capability Map Report, 2025 Assessment (Sept 26 2025)"
AS_OF = "2025-09-26"
REGION_SOURCE_DOC = "AESO ISO Tariff Section 7, effective 2026-01-01"
REGION_AS_OF = "2026-01-01"

PLANNING_AREAS: dict[int, tuple[str, str]] = {
    17: ("Rainbow Lake", "Northwest"), 18: ("High Level", "Northwest"),
    19: ("Peace River", "Northwest"), 20: ("Grande Prairie", "Northwest"),
    21: ("High Prairie", "Northwest"), 22: ("Grande Cache", "Northwest"),
    23: ("Valleyview", "Northwest"), 24: ("Fox Creek", "Northwest"),
    26: ("Swan Hills", "Northwest"), 25: ("Fort McMurray", "Northeast"),
    27: ("Athabasca/Lac La Biche", "Northeast"), 33: ("Fort Saskatchewan", "Northeast"),
    31: ("Wetaskiwin", "Edmonton"), 40: ("Lake Wabamun", "Edmonton"),
    60: ("Edmonton", "Edmonton"), 13: ("Lloydminster", "Central"),
    28: ("Cold Lake", "Central"), 29: ("Hinton/Edson", "Central"),
    30: ("Drayton Valley", "Central"), 32: ("Wainwright", "Central"),
    34: ("Abraham Lake", "Central"), 35: ("Red Deer", "Central"),
    36: ("Alliance/Battle River", "Central"), 37: ("Provost", "Central"),
    38: ("Caroline", "Central"), 39: ("Didsbury", "Central"),
    42: ("Hanna", "Central"), 56: ("Vegreville", "Central"),
    6: ("Calgary", "Calgary"), 57: ("Airdrie", "Calgary"),
    4: ("Medicine Hat", "South"), 43: ("Sheerness", "South"),
    44: ("Seebe", "South"), 45: ("Strathmore/Blackie", "South"),
    46: ("High River", "South"), 47: ("Brooks", "South"),
    48: ("Empress", "South"), 49: ("Stavely", "South"),
    52: ("Vauxhall", "South"), 53: ("Fort Macleod", "South"),
    54: ("Lethbridge", "South"), 55: ("Glenwood", "South"),
}

ATTACHMENT_C_ASSETS = [
    ("VBN1", "Benalto 1", 5, 35, "Red Deer", "Central"),
    ("ACD1", "Big Sky Solar", 140, 48, "Empress", "South"),
    ("VBR1", "Briker 1", 5, 13, "Lloydminster", "Central"),
    ("BPW1", "Buffalo Plains", 466, 49, "Stavely", "South"),
    ("FRM1", "Forty Mile Bow Island", 266, 4, "Medicine Hat", "South"),
    ("FMG1", "Forty Mile Granlea", 20, 4, "Medicine Hat", "South"),
    ("FCS1", "Fox Coulee Solar", 80, 42, "Hanna", "Central"),
    ("GNR1", "Genesee Repower 1", 66, 40, "Wabamun", "Edmonton"),
    ("GNR2", "Genesee Repower 2", 66, 40, "Wabamun", "Edmonton"),
    ("HAL2", "Halkirk 2", 122, 36, "Alliance/Battle River", "Central"),
    ("KH3", "Keephills", 3, 40, "Wabamun", "Edmonton"),
    ("VKW1", "Kenilworth 1", 5, 13, "Lloydminster", "Central"),
    ("CLD1", "Lethbridge Solar", 9, 54, "Lethbridge", "South"),
    ("VNT1", "Netook 1", 5, 39, "Didsbury", "Central"),
    ("VNV1", "Northern Valley 1", 5, 13, "Lloydminster", "Central"),
    ("SCR1", "Suncor Base Plant", 856, 25, "Fort McMurray", "Northeast"),
    ("WPT1", "Wapiti Power Plant", 30, 20, "Grande Prairie", "Northwest"),
    ("WIR1", "Wild Rose", 192, 4, "Medicine Hat", "South"),
    ("WIN1", "Winnifred Wind", 136, 4, "Medicine Hat", "South"),
]

# Column left-edges read off the rendered header row (see docstring).
COLS_A = [("name", 60), ("facility_code", 140), ("tfo", 205),
          ("planning_area", 275), ("bus_number", 345),
          ("voltage_kv", 415), ("capability_mw", 485)]
COLS_B = [("line_name", 60), ("voltage_kv", 160), ("substation_name", 215),
          ("facility_code", 290), ("planning_area", 355),
          ("tfo", 420), ("capability_mw", 485)]
COLS_C = [("asset", 60), ("capability_change_mw", 190),
          ("planning_area", 310), ("region", 430)]


def cluster_lines(words, tol=3.0):
    """Group words into physical lines by vertical position."""
    buckets = defaultdict(list)
    for w in words:
        placed = False
        for key in buckets:
            if abs(key - w["top"]) <= tol:
                buckets[key].append(w)
                placed = True
                break
        if not placed:
            buckets[w["top"]].append(w)
    return [sorted(ws, key=lambda x: x["x0"]) for _, ws in sorted(buckets.items())]


def to_columns(line_words, cols):
    """Assign each word to a column by x0. Returns {col_name: [fragments]}."""
    out = {c[0]: [] for c in cols}
    for w in line_words:
        col = cols[0][0]
        for name, x in cols:
            if w["x0"] >= x - 6:
                col = name
        out[col].append(w["text"])
    return out


def area_code(fragments: list[str]) -> Optional[int]:
    """
    Authoritative key for a planning area: the integer before the hyphen.
    Robust to the mid-word wrapping that mangles the area NAME.
    """
    joined = "".join(fragments)
    m = re.search(r"(\d{1,2})\s*-", joined)
    return int(m.group(1)) if m else None


def area_name(fragments: list[str]) -> str:
    """Display-only. Concatenated without spaces then lightly re-spaced."""
    j = "".join(fragments)
    j = re.sub(r"^\d{1,2}-", "", j)
    j = re.sub(r"(?<=[a-z])(?=[A-Z])", " ", j)   # BattleRiver -> Battle River
    return j.strip()


def num(fragments: list[str]) -> Optional[float]:
    for f in fragments:
        m = re.fullmatch(r"-?\d+(?:\.\d+)?", f.replace(",", ""))
        if m:
            return float(m.group())
    return None


def parse_attachment(pdf, first: int, last: int, cols, is_record) -> list[dict]:
    """
    Walk pages, rebuild multi-line records. A 'record line' is one that
    satisfies is_record(); adjacent non-record lines are merged into it as
    continuation fragments.
    """
    records: list[dict] = []
    for pno in range(first, last + 1):
        lines = cluster_lines(pdf.pages[pno - 1].extract_words())
        parsed = [to_columns(l, cols) for l in lines]
        flags = [is_record(p) for p in parsed]
        rec_idx = [i for i, f in enumerate(flags) if f]
        if not rec_idx:
            continue

        # Assign every continuation line to exactly ONE record — the nearest,
        # and only when directly adjacent. Absorbing both neighbours (the first
        # version) let one continuation bleed into two records: it paired asset
        # GNR2 with "Halkirk 2" and produced planning areas like
        # "Sheerness47-Brooks36-". A continuation belongs to one row only.
        owner: dict[int, int] = {}
        for j, f in enumerate(flags):
            if f:
                continue
            near = min(rec_idx, key=lambda r: (abs(r - j), r))
            if abs(near - j) == 1:
                owner[j] = near

        for i in rec_idx:
            merged = {k: list(v) for k, v in parsed[i].items()}
            for j in sorted(k for k, v in owner.items() if v == i):
                for k, v in parsed[j].items():
                    if v:
                        merged[k] = (merged[k] + v) if j > i else (v + merged[k])
            records.append(merged)
    return records


def rec_a(p):  # substation: needs facility code + numeric bus + voltage
    fc = "".join(p["facility_code"])
    return bool(re.fullmatch(r"\d{2,4}[SPT]", fc)) and num(p["bus_number"]) is not None \
        and num(p["voltage_kv"]) is not None


def rec_b(p):  # line: needs facility code + voltage
    fc = "".join(p["facility_code"])
    return bool(re.fullmatch(r"\d{2,4}[SPT]", fc)) and num(p["voltage_kv"]) is not None


def rec_c(p):  # asset: needs an (ASSET_ID) and a region word
    return bool(re.search(r"\([A-Z0-9]{3,8}\)", " ".join(p["asset"]))) and bool(p["region"])


def split_area(value) -> tuple[Optional[int], str]:
    text = str(value or "").strip()
    match = re.match(r"(\d{1,2})\s*-\s*(.*)", text)
    return (int(match.group(1)), match.group(2).strip()) if match else (None, text)


def extract():
    if not os.path.exists(WORKBOOK):
        log.error("Workbook not found: %s (set AESO_CAPABILITY_XLSX)", WORKBOOK)
        sys.exit(1)

    workbook = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    sub_sheet = workbook["Substation Capabilities Table"]
    line_sheet = workbook["Line Capabilities Table"]

    subs = []
    for row in sub_sheet.iter_rows(min_row=3, values_only=True):
        facility_name, facility_code, tfo, area = row[:4]
        area_id, area_label = split_area(area)
        for start in (4, 7, 10, 13):
            bus, voltage, capability = row[start:start + 3]
            if not isinstance(bus, (int, float)) or not isinstance(capability, (int, float)):
                continue
            subs.append({
                "facility_name": str(facility_name or "").strip(),
                "facility_code": str(facility_code or "").strip(),
                "tfo": str(tfo or "").strip(),
                "planning_area_code": area_id,
                "planning_area_name": area_label,
                "bus_number": int(bus),
                "voltage_kv": int(voltage),
                "capability_mw": float(capability),
            })

    lines = []
    for row in line_sheet.iter_rows(min_row=2, values_only=True):
        line_name, voltage, substation_name, facility_code, area, tfo, capability = row[:7]
        if not line_name or not isinstance(capability, (int, float)):
            continue
        area_id, area_label = split_area(area)
        lines.append({
            "line_name": str(line_name).strip(),
            "voltage_kv": int(voltage),
            "substation_name": str(substation_name or "").strip(),
            "facility_code": str(facility_code or "").strip(),
            "planning_area_code": area_id,
            "planning_area_name": area_label,
            "tfo": str(tfo or "").strip(),
            "capability_mw": float(capability),
        })

    assets = [{
        "asset_id": asset_id,
        "asset_name": asset_name,
        "capability_change_mw": float(change_mw),
        "planning_area_code": area_id,
        "planning_area_name": area_label,
        "region": region,
    } for asset_id, asset_name, change_mw, area_id, area_label, region
      in ATTACHMENT_C_ASSETS]

    workbook.close()
    return subs, lines, assets


def inspect(subs, lines, assets):
    log.info("=== Attachment A — Substation Capability: %d rows ===", len(subs))
    for r in subs[:12]:
        log.info("  %-22s %-6s %-9s area=%-3s %-22s bus=%-7s %3dkV  %s MW",
                 r["facility_name"][:22], r["facility_code"], r["tfo"][:9],
                 r["planning_area_code"], r["planning_area_name"][:22],
                 r["bus_number"], r["voltage_kv"], r["capability_mw"])

    log.info("\n=== Attachment B — Line Capability: %d rows ===", len(lines))
    for r in lines[:12]:
        log.info("  %-24s %3dkV %-20s %-6s area=%-3s %-9s %s MW",
                 r["line_name"][:24], r["voltage_kv"], r["substation_name"][:20],
                 r["facility_code"], r["planning_area_code"], r["tfo"][:9],
                 r["capability_mw"])

    log.info("\n=== Attachment C — 2024 Generation Assets: %d rows ===", len(assets))
    for r in assets:
        log.info("  %-10s %-26s %6s MW  area=%-3s %-22s region=%s",
                 r["asset_id"], r["asset_name"][:26], r["capability_change_mw"],
                 r["planning_area_code"], r["planning_area_name"][:22], r["region"])

    log.info("\n=== Planning area -> region (%d, ISO Tariff Section 7) ===", len(PLANNING_AREAS))
    for k, (name, region) in sorted(PLANNING_AREAS.items()):
        log.info("    %-4s %-26s -> %s", k, name, region)

    areas = sorted({r["planning_area_code"] for r in subs if r["planning_area_code"]})
    log.info("\n=== Planning areas seen in Attachment A: %d ===", len(areas))
    log.info("    %s", ", ".join(str(a) for a in areas))
    unmapped_areas = sorted(set(areas) - set(PLANNING_AREAS))
    log.info("    unmapped in current tariff: %s", unmapped_areas or "none")

    bad = [r for r in subs if r["capability_mw"] is None]
    log.info("\n=== Quality ===")
    log.info("  substations missing capability: %d", len(bad))
    log.info("  lines missing capability:       %d",
             len([r for r in lines if r["capability_mw"] is None]))
    log.info("  substation capability MW range: %s .. %s",
             min((r["capability_mw"] for r in subs if r["capability_mw"] is not None), default=None),
             max((r["capability_mw"] for r in subs if r["capability_mw"] is not None), default=None))
    log.info("  distinct bus numbers:           %d", len({r["bus_number"] for r in subs}))
    capability_total = int(sum(r["capability_change_mw"] for r in assets))
    log.info("  Attachment C capability total:  %s MW", capability_total)

    expected = (239, 491, 19, 2477)
    actual = (len(subs), len(lines), len(assets), capability_total)
    if actual != expected or unmapped_areas:
        raise ValueError(
            "source verification failed: "
            f"expected {expected} + no unmapped areas, got {actual}, {unmapped_areas}"
        )
    log.info("  VERIFIED expected counts and Attachment C total.")


def write(subs, lines, assets):
    import psycopg2, psycopg2.extras
    dsn = os.environ.get("DATABASE_URL")
    if not dsn:
        log.error("DATABASE_URL not set")
        sys.exit(1)
    conn = psycopg2.connect(dsn)
    with conn.cursor() as cur:
        cur.execute("""
        CREATE TABLE IF NOT EXISTS aeso_substation_capability (
            id serial PRIMARY KEY,
            facility_name text, facility_code text, tfo text,
            planning_area_code int, planning_area_name text,
            bus_number int, voltage_kv int, capability_mw numeric(10,2),
            source_document text, as_of_date date,
            UNIQUE (facility_code, bus_number, voltage_kv, as_of_date));
        CREATE INDEX IF NOT EXISTS aeso_subcap_area_idx ON aeso_substation_capability (planning_area_code);

        CREATE TABLE IF NOT EXISTS aeso_line_capability (
            id serial PRIMARY KEY,
            line_name text, voltage_kv int, substation_name text, facility_code text,
            planning_area_code int, planning_area_name text, tfo text,
            capability_mw numeric(10,2),
            source_document text, as_of_date date,
            UNIQUE (line_name, facility_code, as_of_date));
        CREATE INDEX IF NOT EXISTS aeso_linecap_area_idx ON aeso_line_capability (planning_area_code);

        CREATE TABLE IF NOT EXISTS aeso_asset_area (
            id serial PRIMARY KEY,
            asset_id text, asset_name text, capability_change_mw numeric(10,2),
            planning_area_code int, planning_area_name text, region text,
            source_document text, as_of_date date,
            UNIQUE (asset_id, as_of_date));

        CREATE TABLE IF NOT EXISTS aeso_planning_areas (
            id serial PRIMARY KEY,
            planning_area_code int NOT NULL,
            planning_area_name text NOT NULL,
            region text NOT NULL,
            source_document text NOT NULL,
            as_of_date date NOT NULL,
            UNIQUE (planning_area_code, as_of_date));
        CREATE INDEX IF NOT EXISTS aeso_planning_area_region_idx ON aeso_planning_areas (region);
        """)
        psycopg2.extras.execute_batch(cur, """
            INSERT INTO aeso_substation_capability
              (facility_name, facility_code, tfo, planning_area_code, planning_area_name,
               bus_number, voltage_kv, capability_mw, source_document, as_of_date)
            VALUES (%(facility_name)s,%(facility_code)s,%(tfo)s,%(planning_area_code)s,
                    %(planning_area_name)s,%(bus_number)s,%(voltage_kv)s,%(capability_mw)s,
                    %(src)s,%(asof)s)
            ON CONFLICT (facility_code, bus_number, voltage_kv, as_of_date) DO UPDATE SET
              capability_mw = EXCLUDED.capability_mw
        """, [{**r, "src": SOURCE_DOC, "asof": AS_OF} for r in subs])

        psycopg2.extras.execute_batch(cur, """
            INSERT INTO aeso_line_capability
              (line_name, voltage_kv, substation_name, facility_code, planning_area_code,
               planning_area_name, tfo, capability_mw, source_document, as_of_date)
            VALUES (%(line_name)s,%(voltage_kv)s,%(substation_name)s,%(facility_code)s,
                    %(planning_area_code)s,%(planning_area_name)s,%(tfo)s,%(capability_mw)s,
                    %(src)s,%(asof)s)
            ON CONFLICT (line_name, facility_code, as_of_date) DO UPDATE SET
              capability_mw = EXCLUDED.capability_mw
        """, [{**r, "src": SOURCE_DOC, "asof": AS_OF} for r in lines])

        psycopg2.extras.execute_batch(cur, """
            INSERT INTO aeso_asset_area
              (asset_id, asset_name, capability_change_mw, planning_area_code,
               planning_area_name, region, source_document, as_of_date)
            VALUES (%(asset_id)s,%(asset_name)s,%(capability_change_mw)s,
                    %(planning_area_code)s,%(planning_area_name)s,%(region)s,%(src)s,%(asof)s)
            ON CONFLICT (asset_id, as_of_date) DO UPDATE SET
              region = EXCLUDED.region
        """, [{**r, "src": SOURCE_DOC, "asof": AS_OF} for r in assets if r["asset_id"]])

        psycopg2.extras.execute_batch(cur, """
            INSERT INTO aeso_planning_areas
              (planning_area_code, planning_area_name, region, source_document, as_of_date)
            VALUES (%(planning_area_code)s, %(planning_area_name)s, %(region)s, %(src)s, %(asof)s)
            ON CONFLICT (planning_area_code, as_of_date) DO UPDATE SET
              planning_area_name = EXCLUDED.planning_area_name,
              region = EXCLUDED.region,
              source_document = EXCLUDED.source_document
        """, [{
            "planning_area_code": code, "planning_area_name": name,
            "region": region, "src": REGION_SOURCE_DOC, "asof": REGION_AS_OF,
        } for code, (name, region) in PLANNING_AREAS.items()])
    conn.commit()
    conn.close()
    log.info("Wrote %d substations, %d lines, %d assets.", len(subs), len(lines), len(assets))


if __name__ == "__main__":
    s, l, a = extract()
    if "--write" in sys.argv:
        inspect(s, l, a)
        write(s, l, a)
    else:
        inspect(s, l, a)
        log.info("\n(read-only — pass --write to load into the database)")
